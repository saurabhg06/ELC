from flask import Flask, render_template, request, send_file, redirect, url_for, jsonify
import os
from utils.excel_handler import (
    save_to_excel, get_projects, get_project_data, update_project_data, 
    save_area_statement, get_area_statements, save_machine_load, get_machine_loads,
    save_hvac_load, get_hvac_loads  # These two should be defined and imported
)
from openpyxl import load_workbook

app = Flask(__name__)

EXCEL_FILE_PATH = 'mep_load_calculation.xlsx'

@app.route('/')
def index():
    # Get list of existing projects from Excel file
    projects = get_projects(EXCEL_FILE_PATH)
    return render_template('index.html', projects=projects)

@app.route('/results')
def results():
    # Get list of existing projects from Excel file
    projects = get_projects(EXCEL_FILE_PATH)
    return render_template('results.html', projects=projects)

@app.route('/get_project_data', methods=['GET'])
def get_project():
    project_name = request.args.get('project_name')
    sheet_name = request.args.get('sheet_name')
    
    if not project_name or not sheet_name:
        return jsonify({"error": "Missing project name or sheet name"}), 400
    
    project_data = get_project_data(EXCEL_FILE_PATH, sheet_name, project_name)
    if project_data:
        return jsonify(project_data)
    else:
        return jsonify({"error": "Project not found"}), 404

@app.route('/calculate', methods=['POST'])
def calculate():
    try:
        # Get user inputs from the form
        user_inputs = request.form.to_dict()
        
        # Extract project information
        sheet_name = user_inputs.get('sheet_name', 'General')
        project_name = user_inputs.get('project_name', 'Unnamed Project')
        is_update = user_inputs.get('is_update', 'false') == 'true'
        original_project_name = user_inputs.get('original_project_name', '')
        
        # Convert numeric inputs to appropriate types
        try:
            # Convert area to float
            area = float(user_inputs.get('area', 0))
            user_inputs['area'] = area
            
            # Convert load values to float
            equipment_load = float(user_inputs.get('equipment_load', 0))
            lighting_load = float(user_inputs.get('lighting_load', 0))
            hvac_load = float(user_inputs.get('hvac_load', 0))
            safety_factor = float(user_inputs.get('safety_factor', 1.2))
            
            user_inputs['equipment_load'] = equipment_load
            user_inputs['lighting_load'] = lighting_load
            user_inputs['hvac_load'] = hvac_load
            user_inputs['safety_factor'] = safety_factor
            
        except ValueError as e:
            return jsonify({
                'success': False,
                'message': f"Invalid numeric input: {str(e)}"
            }), 400
        
        # Perform calculations
        total_load = equipment_load + lighting_load + hvac_load
        adjusted_load = total_load * safety_factor
        
        # Calculate loads per square foot if area > 0
        if area > 0:
            load_per_sqft = total_load / area
        else:
            load_per_sqft = 0
            
        # Create results dictionary
        results = {
            "Total Load (kW)": round(total_load, 2),
            "Adjusted Load (kW)": round(adjusted_load, 2),
            "Load Density (W/sqft)": round(load_per_sqft * 1000, 2) if area > 0 else "N/A"
        }
        
        # Save to Excel based on whether it's an update or new project
        if is_update and original_project_name:
            filename = update_project_data(
                EXCEL_FILE_PATH, 
                sheet_name, 
                original_project_name,
                user_inputs,
                results
            )
        else:
            filename = save_to_excel(
                user_inputs, 
                results, 
                filename=EXCEL_FILE_PATH, 
                sheet_name=sheet_name
            )
        
        return jsonify({
            'success': True,
            'file_url': f'/download/{filename}'
        })
    except Exception as e:
        import traceback
        print(traceback.format_exc())  # This will print detailed error info to console
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/download/<path:filename>')
def download_file(filename):
    # Get the directory of the app.py file
    directory = os.path.dirname(os.path.abspath(__file__))
    
    # Extract just the filename without directory path
    base_filename = os.path.basename(filename)
    
    return send_file(
        os.path.join(directory, base_filename),
        as_attachment=True, 
        download_name=base_filename
    )

@app.route('/area-statement')
def area_statement():
    """Render area statement page"""
    projects = get_area_statements(EXCEL_FILE_PATH)
    return render_template('area_statement.html', projects=projects)

@app.route('/save-area-statement', methods=['POST'])
def save_area_data():
    """Save area statement data to Excel"""
    try:
        # Get the form data
        data = request.get_json()
        
        project_name = data.get('project_name', 'Unnamed Project')
        areas = data.get('areas', [])
        
        # Validate input
        if not project_name:
            return jsonify({'success': False, 'message': 'Project name is required'}), 400
        
        if not areas or len(areas) == 0:
            return jsonify({'success': False, 'message': 'No area data provided'}), 400
        
        # Save to Excel
        filename = save_area_statement(
            project_name,
            areas,
            filename=EXCEL_FILE_PATH
        )
        
        return jsonify({
            'success': True,
            'file_url': f'/download/{filename}',
            'message': 'Area statement saved successfully!'
        })
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/get-area-data', methods=['GET'])
def get_area_data():
    """Get area statement data for a specific project"""
    project_name = request.args.get('project_name')
    
    if not project_name:
        return jsonify({'success': False, 'message': 'Project name is required'}), 400
    
    try:
        # Load workbook
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet_name = f"Area - {project_name}"
            
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract data for the specific project
                project_data = []
                
                # Start from row 4 (data begins after header)
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, values_only=True):
                    # If we've reached a row with no serial number, we've hit the summary section
                    if not row[0] or not isinstance(row[0], (int, float)):
                        break
                    
                    project_data.append({
                        'floor': row[1],
                        'description': row[2],
                        'area': row[3]
                    })
                
                return jsonify({
                    'success': True,
                    'project_name': project_name,
                    'areas': project_data
                })
        
        return jsonify({
            'success': False,
            'message': 'Project not found'
        }), 404
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/machine-load')
def machine_load():
    """Render machine load page"""
    projects = get_machine_loads(EXCEL_FILE_PATH)
    return render_template('machine_load.html', projects=projects)

@app.route('/save-machine-load', methods=['POST'])
def save_machine_data():
    """Save machine load data to Excel"""
    try:
        # Get the form data
        data = request.get_json()
        
        project_name = data.get('project_name', 'Unnamed Project')
        machines = data.get('machines', [])
        
        # Validate input
        if not project_name:
            return jsonify({'success': False, 'message': 'Project name is required'}), 400
        
        if not machines or len(machines) == 0:
            return jsonify({'success': False, 'message': 'No machine data provided'}), 400
        
        # Save to Excel
        filename = save_machine_load(
            project_name,
            machines,
            filename=EXCEL_FILE_PATH
        )
        
        return jsonify({
            'success': True,
            'file_url': f'/download/{filename}',
            'message': 'Machine load data saved successfully!'
        })
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/get-machine-data', methods=['GET'])
def get_machine_data():
    """Get machine load data for a specific project"""
    project_name = request.args.get('project_name')
    
    if not project_name:
        return jsonify({'success': False, 'message': 'Project name is required'}), 400
    
    try:
        # Load workbook
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet_name = f"Machine - {project_name}"
            
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract data for the specific project
                machine_data = []
                current_floor = None
                
                # Start from row 4 (data begins after header)
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, values_only=True):
                    # Skip rows that don't have complete data
                    if not row[0] and row[1]:  # This might be a floor heading
                        continue
                        
                    if not row[0]:  # Skip empty or summary rows
                        continue
                    
                    machine_data.append({
                        'sr_no': row[0],
                        'floor': row[1],
                        'description': row[2],
                        'qty': row[3],
                        'wattage': row[4],
                        'df': row[5],
                        'voltage': row[6],
                        'frequency': row[7],
                        'phase': row[8],
                        'power_backup': row[9]
                    })
                
                return jsonify({
                    'success': True,
                    'project_name': project_name,
                    'machines': machine_data
                })
        
        return jsonify({
            'success': False,
            'message': 'Project not found'
        }), 404
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/hvac-load')
def hvac_load():
    """Render HVAC load page"""
    projects = get_hvac_loads(EXCEL_FILE_PATH)
    return render_template('hvac_load.html', projects=projects)

@app.route('/save-hvac-load', methods=['POST'])
def save_hvac_data():
    """Save HVAC load data to Excel"""
    try:
        # Get the form data
        data = request.get_json()
        
        project_name = data.get('project_name', 'Unnamed Project')
        hvac_units = data.get('hvac_units', [])
        
        # Validate input
        if not project_name:
            return jsonify({'success': False, 'message': 'Project name is required'}), 400
        
        if not hvac_units or len(hvac_units) == 0:
            return jsonify({'success': False, 'message': 'No HVAC unit data provided'}), 400
        
        # Save to Excel
        filename = save_hvac_load(
            project_name,
            hvac_units,
            filename=EXCEL_FILE_PATH
        )
        
        return jsonify({
            'success': True,
            'file_url': f'/download/{filename}',
            'message': 'HVAC load data saved successfully!'
        })
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/get-hvac-data', methods=['GET'])
def get_hvac_data():
    """Get HVAC load data for a specific project"""
    project_name = request.args.get('project_name')
    
    if not project_name:
        return jsonify({'success': False, 'message': 'Project name is required'}), 400
    
    try:
        # Load workbook
        if os.path.exists(EXCEL_FILE_PATH):
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet_name = f"HVAC - {project_name}"
            
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract data for the specific project
                hvac_data = []
                
                # Start from row 4 (data begins after header)
                for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, values_only=True):
                    # Skip rows that don't have complete data
                    if not row[0] and row[1]:  # This might be a floor heading
                        continue
                        
                    if not row[0]:  # Skip empty or summary rows
                        continue
                    
                    hvac_data.append({
                        'sr_no': row[0],
                        'floor': current_floor if 'current_floor' in locals() else '',
                        'description': row[1],
                        'qty': row[2],
                        'wattage': row[7],
                        'df': row[9],
                        'voltage': row[3],
                        'frequency': row[4],
                        'phase': row[5],
                        'power_backup': row[6],
                        'remarks': row[11] if len(row) > 11 else ''
                    })
                    
                    # Update current floor tracking
                    if row[1] and isinstance(row[1], str) and ' Block' in row[1]:
                        current_floor = row[1].replace(' Block', '')
                
                return jsonify({
                    'success': True,
                    'project_name': project_name,
                    'hvac_units': hvac_data
                })
        
        return jsonify({
            'success': False,
            'message': 'Project not found'
        }), 404
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

if __name__ == '__main__':
    app.run(debug=True)