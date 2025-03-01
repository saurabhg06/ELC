from openpyxl import Workbook, load_workbook
import os
import datetime
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def create_excel_file(data, filename='mep_load_calculation.xlsx'):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MEP Load Calculation"
    # Write headers
    headers = ["Input Parameter", "Value"]
    sheet.append(headers)
    # Write data
    for parameter, value in data.items():
        sheet.append([parameter, value])
    # Save the workbook
    workbook.save(filename)
    return filename

def get_projects(filename='mep_load_calculation.xlsx'):
    """Get list of all projects from Excel file with their sheet names"""
    projects = {}
    if not os.path.exists(filename):
        return projects
    try:
        workbook = load_workbook(filename)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # Skip sheet if it doesn't have the expected headers
            if sheet.max_row < 2:
                continue
            # Get all project names in this sheet (from second column)
            project_names = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > 1 and row[1]:  # Project name is in the second column
                    project_names.append(row[1])
            if project_names:
                projects[sheet_name] = project_names
    except Exception as e:
        print(f"Error reading Excel file: {e}")
    return projects

def get_project_data(filename, sheet_name, project_name):
    """Get data for a specific project from Excel file"""
    if not os.path.exists(filename):
        return None
    try:
        # Use pandas to easily read the Excel file
        df = pd.read_excel(filename, sheet_name=sheet_name)
        # Find the row with the specified project name
        project_row = df[df["Project Name"] == project_name]
        if project_row.empty:
            return None
        # Convert to dictionary
        project_data = project_row.iloc[0].to_dict()
        return project_data
    except Exception as e:
        print(f"Error getting project data: {e}")
        return None

def update_project_data(filename, sheet_name, original_project_name, inputs, results):
    """Update existing project data in Excel file"""
    if not os.path.exists(filename):
        return save_to_excel(inputs, results, filename, sheet_name)
    try:
        workbook = load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            return save_to_excel(inputs, results, filename, sheet_name)
        sheet = workbook[sheet_name]
        # Find row with the original project name
        row_to_update = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            if len(row) > 1 and row[1] == original_project_name:
                row_to_update = row_idx
                break
        if not row_to_update:
            return save_to_excel(inputs, results, filename, sheet_name)
        # Update the row with new data
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        project_name = inputs.get('project_name', original_project_name)
        area = inputs.get('area', 0)
        equipment_load = inputs.get('equipment_load', 0)
        lighting_load = inputs.get('lighting_load', 0)
        hvac_load = inputs.get('hvac_load', 0)
        safety_factor = inputs.get('safety_factor', 1.2)
        total_load = results.get('Total Load', 0)
        adjusted_load = results.get('Adjusted Load (with safety factor)', 0)
        # Update the cells in the row
        sheet.cell(row=row_to_update, column=1).value = timestamp
        sheet.cell(row=row_to_update, column=2).value = project_name
        sheet.cell(row=row_to_update, column=3).value = area
        sheet.cell(row=row_to_update, column=4).value = equipment_load
        sheet.cell(row=row_to_update, column=5).value = lighting_load
        sheet.cell(row=row_to_update, column=6).value = hvac_load
        sheet.cell(row=row_to_update, column=7).value = safety_factor
        sheet.cell(row=row_to_update, column=8).value = total_load
        sheet.cell(row=row_to_update, column=9).value = adjusted_load
        workbook.save(filename)
        return os.path.abspath(filename)
    except Exception as e:
        print(f"Error updating project: {e}")
        return save_to_excel(inputs, results, filename, sheet_name)

def save_to_excel(inputs, results, filename='mep_load_calculation.xlsx', sheet_name=None):
    """
    Save user inputs and calculation results to Excel file with multiple sheets.
    Args:
        inputs (dict): Dictionary of user input parameters
        results (dict): Dictionary of calculation results
        filename (str): Name of the Excel file to create
        sheet_name (str): Name of the sheet to use or create
    Returns:
        str: Path to the created Excel file
    """
    # Check if file exists and load it, otherwise create new workbook
    if os.path.exists(filename):
        try:
            workbook = load_workbook(filename)
        except:
            workbook = Workbook()
    else:
        workbook = Workbook()
    # Use specified sheet or 'General' as default
    if not sheet_name:
        sheet_name = inputs.get('sheet_name', 'General')
    # Check if sheet already exists, otherwise create it
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        # If first sheet is the default 'Sheet', remove it after creating our new sheet
        is_new_workbook = len(workbook.sheetnames) == 1 and workbook.sheetnames[0] == 'Sheet'
        # Create new sheet
        sheet = workbook.create_sheet(sheet_name)
        # Add headers to the new sheet
        headers = ["Timestamp", "Project Name", "Area (sq.ft)", "Occupancy Type", 
                  "Equipment Load (kW)", "Lighting Load (kW)", "HVAC Load (kW)",  
                  "Safety Factor", "Total Load (kW)", "Adjusted Load (kW)", "Load Density (W/sqft)"]
        sheet.append(headers)
        # If default sheet exists and we created a new sheet, remove the default one
        if is_new_workbook:
            default_sheet = workbook['Sheet']
            workbook.remove(default_sheet)
    # Prepare data row with proper numeric types
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    project_name = inputs.get('project_name', 'Unnamed Project')
    area = float(inputs.get('area', 0))
    occupancy = inputs.get('occupancy', 'N/A')
    equipment_load = float(inputs.get('equipment_load', 0))
    lighting_load = float(inputs.get('lighting_load', 0))
    hvac_load = float(inputs.get('hvac_load', 0))
    safety_factor = float(inputs.get('safety_factor', 1.2))
    total_load = float(results.get('Total Load (kW)', 0))
    adjusted_load = float(results.get('Adjusted Load (kW)', 0))
    load_density = results.get('Load Density (W/sqft)', 'N/A')
    # Add data row
    data_row = [timestamp, project_name, area, occupancy, 
               equipment_load, lighting_load, hvac_load, safety_factor,
               total_load, adjusted_load, load_density]
    sheet.append(data_row)
    # Create directory for files if it doesn't exist
    os_dir = os.path.dirname(os.path.abspath(filename))
    if not os.path.exists(os_dir):
        os.makedirs(os_dir)
    # Save the workbook
    workbook.save(filename)
    return os.path.basename(filename)

import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter  # Add this import

def save_area_statement(project_name, areas, filename='mep_load_calculation.xlsx'):
    """
    Save area statement data to Excel file with calculated loads
    
    Args:
        project_name (str): Name of the project
        areas (list): List of dictionaries containing floor, room description and area
        filename (str): Name of the Excel file
        
    Returns:
        str: Path to the created Excel file
    """
    # Check if file exists and load it, otherwise create new workbook
    if os.path.exists(filename):
        try:
            workbook = load_workbook(filename)
        except:
            workbook = Workbook()
    else:
        workbook = Workbook()
    
    # Create a sheet name based on project name (avoid duplicate names)
    sheet_name = f"Area - {project_name}"
    
    # If sheet exists, remove it to recreate
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    
    # Create new sheet
    sheet = workbook.create_sheet(sheet_name)
    
    # Add project name as title
    sheet.cell(row=1, column=1).value = f"PROJECT: {project_name}"
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.merge_cells('A1:F1')
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Add headers to the sheet with formatting
    headers = ["SR.NO.", "FLOOR", "DESCRIPTION OF ROOM", "AREA (SQ.FT)", "1KVA/100SQ.FT"]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center')
    
    # Track floors for summary
    floors = {}
    total_area = 0
    total_load = 0
    current_row = 4  # Start data from row 4
    
    # Add area data rows
    for i, area_data in enumerate(areas, 1):
        floor = area_data.get('floor', '')
        description = area_data.get('description', '')
        area = float(area_data.get('area', 0))
        
        # Calculate 1KVA/100SQ.FT
        load = area / 100  # Load calculated as 1KVA per 100 sq.ft
        
        # Add data row
        sheet.cell(row=current_row, column=1).value = i
        sheet.cell(row=current_row, column=2).value = floor
        sheet.cell(row=current_row, column=3).value = description
        sheet.cell(row=current_row, column=4).value = area
        sheet.cell(row=current_row, column=5).value = load
        
        # Add border to data cells
        for col in range(1, 6):
            sheet.cell(row=current_row, column=col).border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
        
        current_row += 1
        
        # Track floor totals
        if floor not in floors:
            floors[floor] = {'area': 0, 'load': 0}
        floors[floor]['area'] += area
        floors[floor]['load'] += load
        
        total_area += area
        total_load += load
    
    # Add blank row
    current_row += 1
    
    # Add floor summary header
    summary_row = current_row
    sheet.cell(row=summary_row, column=2).value = "FLOOR SUMMARIES"
    sheet.cell(row=summary_row, column=3).value = "TOTAL"
    sheet.cell(row=summary_row, column=2).font = Font(bold=True)
    sheet.cell(row=summary_row, column=3).font = Font(bold=True)
    sheet.cell(row=summary_row, column=2).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    sheet.cell(row=summary_row, column=3).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Add border to summary header
    for col in range(2, 4):
        sheet.cell(row=summary_row, column=col).border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    current_row += 1
    
    # Add floor summary rows
    for floor, data in floors.items():
        sheet.cell(row=current_row, column=2).value = floor
        sheet.cell(row=current_row, column=4).value = data['area']
        sheet.cell(row=current_row, column=5).value = data['load']
        
        # Add border to floor summary cells
        for col in [2, 4, 5]:
            sheet.cell(row=current_row, column=col).border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
        
        current_row += 1
    
    # Add grand total
    grand_total_row = current_row
    sheet.cell(row=grand_total_row, column=2).value = "GRAND TOTAL"
    sheet.cell(row=grand_total_row, column=4).value = total_area
    sheet.cell(row=grand_total_row, column=5).value = total_load
    
    # Format grand total row
    sheet.cell(row=grand_total_row, column=2).font = Font(bold=True)
    sheet.cell(row=grand_total_row, column=4).font = Font(bold=True)
    sheet.cell(row=grand_total_row, column=5).font = Font(bold=True)
    
    # Add double border to grand total cells
    for col in [2, 4, 5]:
        sheet.cell(row=grand_total_row, column=col).border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='double')
        )
    
    # Create a load summary sheet
    summary_sheet_name = f"Summary - {project_name}"
    if summary_sheet_name in workbook.sheetnames:
        del workbook[summary_sheet_name]
        
    summary_sheet = workbook.create_sheet(summary_sheet_name)
    
    # Add title to summary sheet
    summary_sheet.cell(row=1, column=1).value = f"LOAD SUMMARY - {project_name}"
    summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    summary_sheet.merge_cells('A1:D1')
    summary_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Add headers to summary sheet
    summary_headers = ["SR.NO.", "DESCRIPTION", "AREA (SQ.FT)", "1KVA/100SQ.FT"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = summary_sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center')
    
    # Add floor data to summary sheet
    row_idx = 4
    for i, (floor, data) in enumerate(floors.items(), 1):
        summary_sheet.cell(row=row_idx, column=1).value = i
        summary_sheet.cell(row=row_idx, column=2).value = floor
        summary_sheet.cell(row=row_idx, column=3).value = data['area']
        summary_sheet.cell(row=row_idx, column=4).value = data['load']
        
        # Add border to cells
        for col in range(1, 5):
            summary_sheet.cell(row=row_idx, column=col).border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
        
        row_idx += 1
    
    # Add total row to summary sheet
    summary_sheet.cell(row=row_idx, column=2).value = "TOTAL"
    summary_sheet.cell(row=row_idx, column=3).value = total_area
    summary_sheet.cell(row=row_idx, column=4).value = total_load
    
    # Format total row
    summary_sheet.cell(row=row_idx, column=2).font = Font(bold=True)
    summary_sheet.cell(row=row_idx, column=3).font = Font(bold=True)
    summary_sheet.cell(row=row_idx, column=4).font = Font(bold=True)
    
    # Add double border to total cells
    for col in [2, 3, 4]:
        summary_sheet.cell(row=row_idx, column=col).border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='double')
        )
    
    # Adjust column widths in both sheets
    for sheet_obj in [sheet, summary_sheet]:
        sheet_obj.column_dimensions['A'].width = 10
        sheet_obj.column_dimensions['B'].width = 15
        sheet_obj.column_dimensions['C'].width = 30
        sheet_obj.column_dimensions['D'].width = 15
        sheet_obj.column_dimensions['E'].width = 15
    
    # Save the workbook
    workbook.save(filename)
    return os.path.basename(filename)

def get_area_statements(filename='mep_load_calculation.xlsx'):
    """Get list of all projects with area statements"""
    projects = []
    
    if not os.path.exists(filename):
        return projects
    
    try:
        workbook = load_workbook(filename)
        
        # Look for sheets that start with "Area - "
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("Area - "):
                project_name = sheet_name[7:]  # Remove "Area - " prefix
                projects.append(project_name)
        
        projects.sort()
    except Exception as e:
        print(f"Error reading area statements: {e}")
    
    return projects

def save_machine_load(project_name, machines, filename='mep_load_calculation.xlsx'):
    """
    Save machine load data to Excel file with calculated values
    
    Args:
        project_name (str): Name of the project
        machines (list): List of dictionaries containing machine data
        filename (str): Name of the Excel file
        
    Returns:
        str: Path to the created Excel file
    """
    # Check if file exists and load it, otherwise create new workbook
    if os.path.exists(filename):
        try:
            workbook = load_workbook(filename)
        except:
            workbook = Workbook()
    else:
        workbook = Workbook()
    
    # Create a sheet name based on project name
    sheet_name = f"Machine - {project_name}"
    
    # If sheet exists, remove it to recreate
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    
    # Create new sheet
    sheet = workbook.create_sheet(sheet_name)
    
    # Add project name as title
    sheet.cell(row=1, column=1).value = f"MACHINE LOAD LIST - {project_name}"
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.merge_cells('A1:K1')
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Add headers to the sheet with formatting
    headers = [
        "SR.NO.",  
        "FLOOR", 
        "DESCRIPTION OF M/C", 
        "QTY", 
        "WATTAGE / CONNECTED LOAD",
        "D/F",
        "VOLTAGE (V)",
        "FREQUENCY (Hz)",
        "PHASE (3Ph/1Ph)",
        "POWER BACKUP",
        "TOTAL CONNECTED LOAD (KW)",
        "MAX. DEMAND (KW)"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Adjust column widths for better readability
    column_widths = {
        1: 10,  # SR.NO.
        2: 15,  # FLOOR
        3: 25,  # DESCRIPTION OF M/C
        4: 10,  # QTY
        5: 15,  # WATTAGE / CONNECTED LOAD
        6: 10,  # D/F
        7: 12,  # VOLTAGE
        8: 12,  # FREQUENCY
        9: 15,  # PHASE
        10: 20, # POWER BACKUP
        11: 15, # TOTAL CONNECTED LOAD
        12: 15  # MAX. DEMAND
    }
    
    for col_num, width in column_widths.items():
        if col_num <= len(headers):
            sheet.column_dimensions[chr(64 + col_num)].width = width
    
    # Track floors for summary
    floors = {}
    total_connected_load = 0
    total_max_demand = 0
    current_row = 4  # Start data from row 4
    
    # Group machines by floor
    floor_machines = {}
    for machine in machines:
        floor = machine.get('floor', '')
        if floor not in floor_machines:
            floor_machines[floor] = []
        floor_machines[floor].append(machine)
    
    # Add data by floor
    sr_no = 1
    for floor, floor_data in floor_machines.items():
        # Add floor heading
        floor_heading = f"{floor}"
        sheet.cell(row=current_row, column=2).value = floor_heading
        sheet.cell(row=current_row, column=2).font = Font(bold=True)
        current_row += 1
        
        # Add machine data for this floor
        floor_connected_load = 0
        floor_max_demand = 0
        
        for machine in floor_data:
            description = machine.get('description', '')
            qty = int(machine.get('qty', 0))
            wattage = float(machine.get('wattage', 0))
            df = float(machine.get('df', 1.0))
            voltage = machine.get('voltage', '230')
            frequency = machine.get('frequency', '50')
            phase = machine.get('phase', '1Ph')
            power_backup = machine.get('power_backup', 'MSEB')
            
            # Calculate derived values
            total_load = wattage * qty
            max_demand = total_load * df
            
            # Add data row
            sheet.cell(row=current_row, column=1).value = sr_no
            sheet.cell(row=current_row, column=2).value = floor
            sheet.cell(row=current_row, column=3).value = description
            sheet.cell(row=current_row, column=4).value = qty
            sheet.cell(row=current_row, column=5).value = wattage
            sheet.cell(row=current_row, column=6).value = df
            sheet.cell(row=current_row, column=7).value = voltage
            sheet.cell(row=current_row, column=8).value = frequency
            sheet.cell(row=current_row, column=9).value = phase
            sheet.cell(row=current_row, column=10).value = power_backup
            sheet.cell(row=current_row, column=11).value = total_load
            sheet.cell(row=current_row, column=12).value = max_demand
            
            # Add borders to all cells in the row
            for col in range(1, 13):
                sheet.cell(row=current_row, column=col).border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                sheet.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
            
            # Track totals
            floor_connected_load += total_load
            floor_max_demand += max_demand
            sr_no += 1
            current_row += 1
            
        # Add floor summary row
        sheet.cell(row=current_row, column=2).value = f"{floor} Total"
        sheet.cell(row=current_row, column=2).font = Font(bold=True)
        sheet.cell(row=current_row, column=11).value = floor_connected_load
        sheet.cell(row=current_row, column=12).value = floor_max_demand
        
        # Format floor summary row
        for col in [2, 11, 12]:
            sheet.cell(row=current_row, column=col).border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='double')
            )
            sheet.cell(row=current_row, column=col).font = Font(bold=True)
        
        # Add to totals
        total_connected_load += floor_connected_load
        total_max_demand += floor_max_demand
        
        # Track for summary sheet
        floors[floor] = {
            'connected_load': floor_connected_load,
            'max_demand': floor_max_demand
        }
        
        current_row += 2  # Add some space between floors
    
    # Add grand total row
    grand_total_row = current_row
    sheet.cell(row=grand_total_row, column=2).value = "GRAND TOTAL"
    sheet.cell(row=grand_total_row, column=11).value = total_connected_load
    sheet.cell(row=grand_total_row, column=12).value = total_max_demand
    
    # Format grand total row
    for col in [2, 11, 12]:
        sheet.cell(row=grand_total_row, column=col).font = Font(bold=True)
        sheet.cell(row=grand_total_row, column=col).border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='double'), 
            bottom=Side(style='double')
        )
    
    # Create a load summary sheet
    summary_sheet_name = f"Machine Summary - {project_name}"
    if summary_sheet_name in workbook.sheetnames:
        del workbook[summary_sheet_name]
        
    summary_sheet = workbook.create_sheet(summary_sheet_name)
    
    # Add title to summary sheet
    summary_sheet.cell(row=1, column=1).value = f"MACHINE LOAD SUMMARY - {project_name}"
    summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    summary_sheet.merge_cells('A1:E1')
    summary_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Add headers to summary sheet
    summary_headers = ["SR.NO.", "FLOOR", "CONNECTED LOAD (KW)", "DIVERSITY FACTOR", "MAX. DEMAND (KW)"]
    for col_num, header in enumerate(summary_headers, 1):
        cell = summary_sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths for summary sheet
    summary_sheet.column_dimensions['A'].width = 10
    summary_sheet.column_dimensions['B'].width = 20
    summary_sheet.column_dimensions['C'].width = 20
    summary_sheet.column_dimensions['D'].width = 15
    summary_sheet.column_dimensions['E'].width = 20
    
    # Add floor data to summary sheet
    row_idx = 4
    for i, (floor, data) in enumerate(floors.items(), 1):
        summary_sheet.cell(row=row_idx, column=1).value = i
        summary_sheet.cell(row=row_idx, column=2).value = floor
        summary_sheet.cell(row=row_idx, column=3).value = data['connected_load']
        
        # Calculate average diversity factor or use a default
        df_value = data['max_demand'] / data['connected_load'] if data['connected_load'] > 0 else 0.8
        summary_sheet.cell(row=row_idx, column=4).value = round(df_value, 2)
        
        summary_sheet.cell(row=row_idx, column=5).value = data['max_demand']
        
        # Add border to cells
        for col in range(1, 6):
            summary_sheet.cell(row=row_idx, column=col).border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            summary_sheet.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
        
        row_idx += 1
    
    # Add total row to summary sheet
    summary_sheet.cell(row=row_idx, column=2).value = "TOTAL"
    summary_sheet.cell(row=row_idx, column=3).value = total_connected_load
    
    # Calculate overall diversity factor
    overall_df = total_max_demand / total_connected_load if total_connected_load > 0 else 0.8
    summary_sheet.cell(row=row_idx, column=4).value = round(overall_df, 2)
    
    summary_sheet.cell(row=row_idx, column=5).value = total_max_demand
    
    # Format total row
    for col in [2, 3, 4, 5]:
        summary_sheet.cell(row=row_idx, column=col).font = Font(bold=True)
        summary_sheet.cell(row=row_idx, column=col).border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='double')
        )
        summary_sheet.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
    
    # Save the workbook
    workbook.save(filename)
    return os.path.basename(filename)

def get_machine_loads(filename='mep_load_calculation.xlsx'):
    """Get list of all machine load projects"""
    projects = []
    
    if not os.path.exists(filename):
        return projects
    
    try:
        workbook = load_workbook(filename)
        
        # Look for sheets that start with "Machine - "
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("Machine - "):
                project_name = sheet_name[10:]  # Remove "Machine - " prefix
                projects.append(project_name)
        
        projects.sort()
    except Exception as e:
        print(f"Error reading machine load projects: {e}")
    
    return projects

def save_hvac_load(project_name, hvac_units, filename='mep_load_calculation.xlsx'):
    """
    Save HVAC load data to Excel file with calculated loads
    
    Args:
        project_name (str): Name of the project
        hvac_units (list): List of dictionaries containing HVAC unit data
        filename (str): Name of the Excel file
        
    Returns:
        str: Path to the created Excel file
    """
    # Check if file exists and load it, otherwise create new workbook
    if os.path.exists(filename):
        try:
            workbook = load_workbook(filename)
        except:
            workbook = Workbook()
    else:
        workbook = Workbook()
    
    # Create a sheet name based on project name
    sheet_name = f"HVAC - {project_name}"
    
    # If sheet exists, remove it to recreate
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    
    # Create new sheet
    sheet = workbook.create_sheet(sheet_name)
    
    # Add project name as title
    sheet.cell(row=1, column=1).value = f"PROJECT: {project_name} - HVAC LOAD LIST"
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
    sheet.merge_cells('A1:L1')
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Add headers to the sheet with formatting
    headers = [
        "Sr.No", "DESCRIPTION", "Quantity", "Voltage (V)", "Frequency (Hz)", 
        "Phase (3Ph/1Ph)", "Power Backup (DG/MSEB/UPS/inverter)", 
        "Connected Load (TR)", "Total Connected Load (TR)", 
        "Diversity Factor", "Max. Demand (TR)", "Remarks (Area in sq.ft)"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Track floors for summary
    floors = {}
    current_row = 4  # Start data from row 4
    
    # Sort hvac_units by floor to group them
    sorted_hvac = sorted(hvac_units, key=lambda x: x.get('floor', ''))
    
    # Keep track of current floor for grouping
    current_floor = None
    total_connected_load = 0
    total_demand_load = 0
    
    for i, hvac in enumerate(sorted_hvac, 1):
        floor = hvac.get('floor', '')
        description = hvac.get('description', '')
        qty = int(hvac.get('qty', 1))
        wattage = float(hvac.get('wattage', 0))  # TR (Tons of Refrigeration)
        diversity_factor = float(hvac.get('df', 0.8))
        voltage = hvac.get('voltage', '230')
        frequency = hvac.get('frequency', '50')
        phase = hvac.get('phase', '1Ph')
        power_backup = hvac.get('power_backup', 'MSEB')
        remarks = hvac.get('remarks', '')
        
        # Calculate total connected load and max demand
        total_load = wattage * qty  # Total TR
        max_demand = total_load * diversity_factor  # Max demand in TR
        
        # Add floor header if this is a new floor
        if floor != current_floor:
            # Add a floor header row
            sheet.cell(row=current_row, column=2).value = f"{floor} Block"
            sheet.cell(row=current_row, column=2).font = Font(bold=True)
            sheet.merge_cells(f'B{current_row}:L{current_row}')
            sheet.cell(row=current_row, column=2).alignment = Alignment(horizontal='left')
            sheet.cell(row=current_row, column=2).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            current_row += 1
            current_floor = floor
            
            # Initialize floor tracking
            if floor not in floors:
                floors[floor] = {'connected_load': 0, 'max_demand': 0}
        
        # Add HVAC unit data row
        sheet.cell(row=current_row, column=1).value = i
        sheet.cell(row=current_row, column=2).value = description
        sheet.cell(row=current_row, column=3).value = qty
        sheet.cell(row=current_row, column=4).value = voltage
        sheet.cell(row=current_row, column=5).value = frequency
        sheet.cell(row=current_row, column=6).value = phase
        sheet.cell(row=current_row, column=7).value = power_backup
        sheet.cell(row=current_row, column=8).value = wattage  # Connected load per unit (TR)
        sheet.cell(row=current_row, column=9).value = total_load  # Total connected load (TR)
        sheet.cell(row=current_row, column=10).value = diversity_factor
        sheet.cell(row=current_row, column=11).value = max_demand  # Max demand (TR)
        sheet.cell(row=current_row, column=12).value = remarks
        
        # Add border to data cells
        for col in range(1, 13):
            sheet.cell(row=current_row, column=col).border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            sheet.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
        
        current_row += 1
        
        # Track floor totals
        floors[floor]['connected_load'] += total_load
        floors[floor]['max_demand'] += max_demand
        
        # Track grand totals
        total_connected_load += total_load
        total_demand_load += max_demand
    
    # Add blank row
    current_row += 1
    
    # Add HVAC to Electrical Conversion section
    conversion_row = current_row
    sheet.cell(row=conversion_row, column=1).value = "HVAC TO ELECTRICAL CONVERSION"
    sheet.cell(row=conversion_row, column=1).font = Font(bold=True)
    sheet.merge_cells(f'A{conversion_row}:L{conversion_row}')
    sheet.cell(row=conversion_row, column=1).alignment = Alignment(horizontal='center')
    sheet.cell(row=conversion_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    current_row += 1
    
    # Add conversion note
    # Note: 1 TR = 3.517 kW
    sheet.cell(row=current_row, column=2).value = "1 TR = 3.517 kW"
    sheet.cell(row=current_row, column=2).font = Font(bold=True)
    sheet.cell(row=current_row, column=2).alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Calculate electrical equivalent
    electrical_connected = total_connected_load * 3.517  # Convert TR to kW
    electrical_demand = total_demand_load * 3.517  # Convert TR to kW
    
    sheet.cell(row=current_row, column=2).value = "Total HVAC Connected Load (TR)"
    sheet.cell(row=current_row, column=3).value = total_connected_load
    current_row += 1
    
    sheet.cell(row=current_row, column=2).value = "Total HVAC Max Demand (TR)"
    sheet.cell(row=current_row, column=3).value = total_demand_load
    current_row += 1
    
    sheet.cell(row=current_row, column=2).value = "Total HVAC Connected Load (kW)"
    sheet.cell(row=current_row, column=3).value = electrical_connected
    current_row += 1
    
    sheet.cell(row=current_row, column=2).value = "Total HVAC Max Demand (kW)"
    sheet.cell(row=current_row, column=3).value = electrical_demand
    current_row += 2
    
    # Add floor summary header
    summary_row = current_row
    sheet.cell(row=summary_row, column=1).value = "SUMMARY BY FLOOR"
    sheet.cell(row=summary_row, column=1).font = Font(bold=True)
    sheet.merge_cells(f'A{summary_row}:L{summary_row}')
    sheet.cell(row=summary_row, column=1).alignment = Alignment(horizontal='center')
    sheet.cell(row=summary_row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    current_row += 1
    
    # Add column headers for summary
    summary_headers = ["Floor", "Connected Load (TR)", "Connected Load (kW)", "Max Demand (TR)", "Max Demand (kW)"]
    for col_num, header in enumerate(summary_headers, 2):  # Start from column B
        cell = sheet.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center')
    
    current_row += 1
    
    # Add floor summary rows
    for floor, data in floors.items():
        tr_connected = data['connected_load']
        tr_demand = data['max_demand']
        kw_connected = tr_connected * 3.517
        kw_demand = tr_demand * 3.517
        
        sheet.cell(row=current_row, column=2).value = floor
        sheet.cell(row=current_row, column=3).value = tr_connected
        sheet.cell(row=current_row, column=4).value = kw_connected
        sheet.cell(row=current_row, column=5).value = tr_demand
        sheet.cell(row=current_row, column=6).value = kw_demand
        
        # Add border to floor summary cells
        for col in range(2, 7):
            sheet.cell(row=current_row, column=col).border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            sheet.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
        
        current_row += 1
    
    # Add grand total
    grand_total_row = current_row
    sheet.cell(row=grand_total_row, column=2).value = "GRAND TOTAL"
    sheet.cell(row=grand_total_row, column=3).value = total_connected_load
    sheet.cell(row=grand_total_row, column=4).value = total_connected_load * 3.517
    sheet.cell(row=grand_total_row, column=5).value = total_demand_load
    sheet.cell(row=grand_total_row, column=6).value = total_demand_load * 3.517
    
    # Format grand total row
    sheet.cell(row=grand_total_row, column=2).font = Font(bold=True)
    sheet.cell(row=grand_total_row, column=3).font = Font(bold=True)
    sheet.cell(row=grand_total_row, column=4).font = Font(bold=True)
    sheet.cell(row=grand_total_row, column=5).font = Font(bold=True)
    sheet.cell(row=grand_total_row, column=6).font = Font(bold=True)
    
    # Add double border to grand total cells
    for col in range(2, 7):
        sheet.cell(row=grand_total_row, column=col).border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='double')
        )
        sheet.cell(row=grand_total_row, column=col).alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    column_widths = [10, 25, 10, 15, 15, 15, 25, 15, 15, 15, 15, 25]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width
    
    # Save the workbook
    workbook.save(filename)
    return os.path.basename(filename)

def get_hvac_loads(filename='mep_load_calculation.xlsx'):
    """Get list of all projects with HVAC loads"""
    projects = []
    
    if not os.path.exists(filename):
        return projects
    
    try:
        workbook = load_workbook(filename)
        
        # Look for sheets that start with "HVAC - "
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("HVAC - "):
                project_name = sheet_name[7:]  # Remove "HVAC - " prefix
                projects.append(project_name)
        
        projects.sort()
    except Exception as e:
        print(f"Error reading HVAC load projects: {e}")
    
    return projects